#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成100条问答对人工抽检表（分层抽样，附gold证据原文，下拉判定列）。"""
import json
import random
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

random.seed(42)

qas = [json.loads(l) for l in open("data/qa_seed.jsonl", encoding="utf-8")]
units = [json.loads(l) for l in open("data/corpus/structural.jsonl", encoding="utf-8")]

# 建 (source, page, table_id) -> [units] 索引
idx = defaultdict(list)
for u in units:
    if u.get("table_id"):
        idx[(u["source"], u["page"], u["table_id"])].append(u)

def qa_years(qa):
    m = qa.get("meta", {})
    ys = m.get("years") or ([m["year"]] if m.get("year") else [])
    return [str(y) for y in ys]


def evidence_text(qa):
    """优先展示与问题年份匹配的列（如问2020年就展示"2020年=..."的三元组）。"""
    years = qa_years(qa)
    lines, seen = [], set()
    for ev in qa["evidence"]:
        key = (ev["source"], ev["page"], ev.get("table_id"))
        hits = [u for u in idx.get(key, [])
                if ev.get("row_label", "") and ev["row_label"] in u["text"]]
        # 年份匹配的排前面：只看行标签之后的"列标题=值"段，避免文档前缀年报年份误匹配
        def year_match(u):
            t = u["text"]
            seg = t.split("|", 1)[1] if "|" in t else t
            return any(f"{y}年" in seg for y in years)
        hits.sort(key=lambda u: not year_match(u))
        for u in hits[:2]:
            t = u["text"]
            if t not in seen:
                seen.add(t)
                lines.append(t)
    return "\n".join(lines[:4]) or "（未匹配到三元组，请对照PDF核查）"

# 分层抽样 53/38/9
by_type = defaultdict(list)
for qa in qas:
    by_type[qa["type"]].append(qa)
sample = (random.sample(by_type["extraction"], 53)
          + random.sample(by_type["yoy_compare"], 38)
          + random.sample(by_type["cross_company"], 9))

TYPE_CN = {"extraction": "抽取", "yoy_compare": "同比", "cross_company": "对比"}

wb = Workbook()
ws = wb.active
ws.title = "抽检表"

thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)
yellow = PatternFill("solid", fgColor="FFFF00")
gray = PatternFill("solid", fgColor="EFEFEF")
bold = Font(name="Arial", bold=True, size=11)
normal = Font(name="Arial", size=10)
wrap = Alignment(wrap_text=True, vertical="top")

ws["A1"] = "CFR-NumQA 问答对人工抽检表（100条，分层抽样：抽取53/同比38/对比9，seed=42）"
ws["A1"].font = Font(name="Arial", bold=True, size=13)
ws["A2"] = ("填写说明：只需填黄色三列。答案正确?/证据支持? 用下拉选择 ✓ / ✗ / 存疑；"
            "有问题时在备注列写一句原因。第4行为示例行（不计入统计）。"
            "证据原文列是自动从语料匹配的gold三元组，若显示未匹配请对照原PDF核查。")
ws["A2"].font = Font(name="Arial", size=10, color="666666")
ws["A2"].alignment = wrap
ws.merge_cells("A1:J1")
ws.merge_cells("A2:J2")
ws.row_dimensions[2].height = 30

headers = ["序号", "ID", "题型", "问题", "标准答案", "证据原文（自动匹配）",
           "来源（文件@页）", "答案正确?", "证据支持?", "备注"]
HDR = 3
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=HDR, column=c, value=h)
    cell.font = bold
    cell.fill = gray
    cell.border = thin
    cell.alignment = Alignment(vertical="center", horizontal="center")

# 示例行
example = ["示例", "seed_0001", "抽取", "平安银行2023年度的营业收入是多少？",
           "164,699百万元",
           "【平安银行 2023年年度报告】[表:2.1 关键指标（货币单位：人民币百万元）] "
           "营业收入 | 2023年 = 164,699百万元",
           "000001_平安银行_2023年年度报告.pdf @15", "✓", "✓",
           "答案与证据一致"]
for c, v in enumerate(example, 1):
    cell = ws.cell(row=HDR + 1, column=c, value=v)
    cell.font = Font(name="Arial", size=10, italic=True, color="888888")
    cell.border = thin
    cell.alignment = wrap

START = HDR + 2
for i, qa in enumerate(sample):
    r = START + i
    src = "; ".join(sorted({f"{e['source']} @{e['page']}" for e in qa["evidence"]}))
    row = [i + 1, qa["id"], TYPE_CN[qa["type"]], qa["question"], qa["answer"],
           evidence_text(qa), src, "", "", ""]
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = normal
        cell.border = thin
        cell.alignment = wrap
    for c in (8, 9, 10):
        ws.cell(row=r, column=c).fill = yellow

dv = DataValidation(type="list", formula1='"✓,✗,存疑"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"H{START}:I{START + len(sample) - 1}")

widths = [5, 10, 6, 34, 16, 55, 30, 9, 9, 18]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = f"A{START}"

# 汇总区（公式）
E = START + len(sample) - 1
s = E + 2
labels = [
    ("已完成判定（答案列）", f'=COUNTIF(H{START}:H{E},"✓")+COUNTIF(H{START}:H{E},"✗")+COUNTIF(H{START}:H{E},"存疑")'),
    ("答案正确 ✓", f'=COUNTIF(H{START}:H{E},"✓")'),
    ("答案错误 ✗", f'=COUNTIF(H{START}:H{E},"✗")'),
    ("答案存疑", f'=COUNTIF(H{START}:H{E},"存疑")'),
    ("证据支持 ✓", f'=COUNTIF(I{START}:I{E},"✓")'),
    ("证据不支持 ✗", f'=COUNTIF(I{START}:I{E},"✗")'),
    ("答案正确率(已判定中)", f'=IF(COUNTIF(H{START}:H{E},"✓")+COUNTIF(H{START}:H{E},"✗")=0,"-",COUNTIF(H{START}:H{E},"✓")/(COUNTIF(H{START}:H{E},"✓")+COUNTIF(H{START}:H{E},"✗")))'),
]
ws.cell(row=s - 1, column=7, value="汇总（自动计算）").font = bold
for j, (lab, f) in enumerate(labels):
    ws.cell(row=s + j, column=7, value=lab).font = normal
    fc = ws.cell(row=s + j, column=8, value=f)
    fc.font = normal
    if "正确率" in lab:
        fc.number_format = "0.0%"

wb.save("抽检表_人工校验100条.xlsx")
print("saved, rows:", len(sample))
