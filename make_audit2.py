#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成第二标注者的抽检工作簿（两个页签），用于计算标注一致性。

为什么要两个页签
----------------
首轮抽检（make_audit.py，seed=42 的 100 条）标注结果是 97✓/4存疑 与 101✓/0，
而那 4 条存疑指出的缺陷（比率型指标误标货币单位，全库 42 条）**已经修复**。
也就是说在当前发布版数据上，随机样本的两列判定几乎必然全是 ✓——零方差。
Cohen's kappa 在零方差下无定义，即使勉强算出也会因边缘分布极度偏斜而塌陷
（kappa 悖论：一致率 ~100%，kappa 却接近 0），报出来比不报还难看。

所以分开报两件事：
  主样本  -> 一致率。证明在有代表性的随机样本上两位标注者判定一致。
  难例样本 -> kappa / Gwet's AC1。刻意挑可争议的条目，让系数有真实变异可衡量。
             在难例上测得的一致性是随机样本上的下界，这样报更强而非更弱。

难例怎么挑（只用于选样，标注表里不透露来源，也绝不显示模型答案——
那会锚定标注者，让"独立标注"名存实亡）：
  A. 模型给出 grounded error：检索到了证据却答得与 gold 不同，两者必有一错
  B. 证据跨表或跨页：核验成本最高
  C. gold 答案带货币单位：单位由 caption 回挂而来，最易错标
  D. 比率/每股类指标：正是产生那 42 条缺陷的那一族，拒绝表的边界地带

用法：
    python make_audit2.py                 # 生成 抽检表_第二标注者.xlsx
    python make_audit2.py --hard-size 80
"""
import argparse
import json
import random
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import re

TYPE_CN = {"extraction": "抽取", "yoy_compare": "同比", "cross_company": "对比"}
RATIO_RE = re.compile(r"率|占比|比例|百分|每股|元/股")
MONEY_RE = re.compile(r"(元|万元|亿元|百万元|千元)$")


def load_units_index():
    units = [json.loads(l) for l in
             open("data/corpus/structural.jsonl", encoding="utf-8")]
    idx = defaultdict(list)
    for u in units:
        if u.get("table_id"):
            idx[(u["source"], u["page"], u["table_id"])].append(u)
    return idx


def qa_years(qa):
    m = qa.get("meta", {})
    ys = m.get("years") or ([m["year"]] if m.get("year") else [])
    return [str(y) for y in ys]


def evidence_text(qa, idx):
    """与 make_audit.py 完全一致：优先展示与问题年份匹配的三元组。"""
    years = qa_years(qa)
    lines, seen = [], set()
    for ev in qa["evidence"]:
        key = (ev["source"], ev["page"], ev.get("table_id"))
        hits = [u for u in idx.get(key, [])
                if ev.get("row_label", "") and ev["row_label"] in u["text"]]

        def year_match(u):
            t = u["text"]
            seg = t.split("|", 1)[1] if "|" in t else t
            return any(f"{y}年" in seg for y in years)

        hits.sort(key=lambda u: not year_match(u))
        for u in hits[:2]:
            if u["text"] not in seen:
                seen.add(u["text"])
                lines.append(u["text"])
    return "\n".join(lines[:4]) or "（未匹配到三元组，请对照PDF核查）"


def main_sample(qas):
    """复现首轮那 100 条（seed=42，分层 53/38/9），供两位标注者配对比较。"""
    rng = random.Random(42)
    by = defaultdict(list)
    for q in qas:
        by[q["type"]].append(q)
    return (rng.sample(by["extraction"], 53)
            + rng.sample(by["yoy_compare"], 38)
            + rng.sample(by["cross_company"], 9))


def hard_sample(qas, exclude_ids, size):
    """四个难例层，去重后打散。层内按题型分散，避免某一层被单一题型占满。"""
    by_id = {q["id"]: q for q in qas}
    rng = random.Random(2026)

    gerr = set()
    try:
        from eval_answers import ABSTAIN_RE, is_correct, is_grounded
        for l in open("results/answers_v3_full.jsonl", encoding="utf-8"):
            r = json.loads(l)
            if (not ABSTAIN_RE.search(r["prediction"])
                    and not is_correct(r) and is_grounded(r)):
                gerr.add(r.get("id"))
    except (ImportError, FileNotFoundError):
        pass

    pools = {
        "A": [q for q in qas if q["id"] in gerr],
        "B": [q for q in qas if len(q.get("evidence", [])) > 1],
        "C": [q for q in qas if MONEY_RE.search(str(q.get("answer", "")))],
        "D": [q for q in qas
              if RATIO_RE.search(str(q.get("meta", {}).get("indicator", "")))],
    }
    quota = {"A": 0.375, "B": 0.25, "C": 0.1875, "D": 0.1875}   # 30/20/15/15 @80

    picked, seen = [], set(exclude_ids)
    for k in "ABCD":
        want = round(size * quota[k])
        cands = [q for q in pools[k] if q["id"] not in seen]
        # 题型轮转，防止某层全是同比题
        bytype = defaultdict(list)
        for q in cands:
            bytype[q["type"]].append(q)
        for v in bytype.values():
            rng.shuffle(v)
        order = sorted(bytype, key=lambda t: -len(bytype[t]))
        while want > 0 and any(bytype[t] for t in order):
            for t in order:
                if want == 0:
                    break
                if bytype[t]:
                    q = bytype[t].pop()
                    picked.append((q, k))
                    seen.add(q["id"])
                    want -= 1
    rng.shuffle(picked)          # 打散，标注表里看不出层次来源
    return picked


def build_sheet(ws, rows, idx, title, note):
    thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    yellow = PatternFill("solid", fgColor="FFFF00")
    gray = PatternFill("solid", fgColor="EFEFEF")
    bold = Font(name="Arial", bold=True, size=11)
    normal = Font(name="Arial", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = note
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = wrap
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 44

    headers = ["序号", "ID", "题型", "问题", "标准答案", "证据原文（自动匹配）",
               "来源（文件@页）", "答案正确?", "证据支持?", "备注"]
    HDR = 3
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=HDR, column=c, value=h)
        cell.font = bold
        cell.fill = gray
        cell.border = thin
        cell.alignment = Alignment(vertical="center", horizontal="center")

    example = ["示例", "seed_0001", "抽取", "平安银行2023年度的营业收入是多少？",
               "164,699百万元",
               "【平安银行 2023年年度报告】[表:2.1 关键指标（货币单位：人民币百万元）] "
               "营业收入 | 2023年 = 164,699百万元",
               "000001_平安银行_2023年年度报告.pdf @15", "✓", "✓", "答案与证据一致"]
    for c, v in enumerate(example, 1):
        cell = ws.cell(row=HDR + 1, column=c, value=v)
        cell.font = Font(name="Arial", size=10, italic=True, color="888888")
        cell.border = thin
        cell.alignment = wrap

    START = HDR + 2
    for i, qa in enumerate(rows):
        r = START + i
        src = "; ".join(sorted({f"{e['source']} @{e['page']}" for e in qa["evidence"]}))
        row = [i + 1, qa["id"], TYPE_CN[qa["type"]], qa["question"], qa["answer"],
               evidence_text(qa, idx), src, "", "", ""]
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = normal
            cell.border = thin
            cell.alignment = wrap
        for c in (8, 9, 10):
            ws.cell(row=r, column=c).fill = yellow

    dv = DataValidation(type="list", formula1='"✓,✗,存疑"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H{START}:I{START + len(rows) - 1}")

    for c, w in enumerate([5, 10, 6, 34, 16, 55, 30, 9, 9, 18], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{START}"
    return START, START + len(rows) - 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--hard-size", type=int, default=80)
    ap.add_argument("--recheck", nargs="*",
                    default=["seed_0282", "seed_0518", "seed_0465", "seed_0364"],
                    help="标注者A需复核的条目：首轮标为存疑、其后数据已修复，"
                         "旧标注对不上当前发布版，不复核会把版本差异算成标注分歧")
    args = ap.parse_args()

    qas = [json.loads(l) for l in open("data/qa_seed.jsonl", encoding="utf-8")]
    by_id = {q["id"]: q for q in qas}
    idx = load_units_index()

    main_rows = main_sample(qas)
    hard = hard_sample(qas, {q["id"] for q in main_rows}, args.hard_size)
    hard_rows = [q for q, _ in hard]
    recheck_rows = [by_id[i] for i in args.recheck if i in by_id]

    NOTE = ("填写说明：只需填黄色三列。答案正确?/证据支持? 用下拉选择 ✓ / ✗ / 存疑；"
            "有疑问时在备注列写一句原因。第4行为示例行（不计入统计）。"
            "请独立判定——不要参考另一位标注者的结果，也不要互相商量，否则一致性系数失去意义。"
            "证据原文列是自动从语料匹配的 gold 三元组，若显示未匹配请对照原PDF核查。")
    HARD_NOTE = (NOTE + " 本页刻意挑选了核验成本高或存在分歧风险的条目（证据跨表跨页、"
                 "单位由表名回挂、比率类指标等）。顺序已打散，判定时无需考虑它为何入选。")

    # 标注者A（首轮已标过主样本）：难例 + 4条复核
    wa = Workbook()
    build_sheet(wa.active, hard_rows, idx,
                f"CFR-NumQA 标注一致性 · 标注者A · 难例样本（{len(hard_rows)}条）", HARD_NOTE)
    wa.active.title = "难例样本"
    if recheck_rows:
        ws = wa.create_sheet("复核4条")
        build_sheet(ws, recheck_rows, idx,
                    f"标注者A · 需复核（{len(recheck_rows)}条）",
                    "这几条首轮标为「存疑」，指出的缺陷（比率型指标误标货币单位）其后已在数据集中修复。"
                    "旧标注对应的是修复前的数据，若直接拿来配对，会把版本差异算成标注者分歧。"
                    "请按当前内容重新判定。" + NOTE)
    wa.save("抽检表_标注者A.xlsx")

    # 标注者B（第二标注者，全新）：主样本 + 难例
    wbk = Workbook()
    build_sheet(wbk.active, main_rows, idx,
                "CFR-NumQA 标注一致性 · 标注者B · 主样本（100条，与首轮同一批：分层 抽取53/同比38/对比9，seed=42）",
                NOTE)
    wbk.active.title = "主样本100条"
    build_sheet(wbk.create_sheet("难例样本"), hard_rows, idx,
                f"CFR-NumQA 标注一致性 · 标注者B · 难例样本（{len(hard_rows)}条）", HARD_NOTE)
    wbk.save("抽检表_标注者B.xlsx")

    from collections import Counter
    print("已生成两份工作簿：")
    print(f"  抽检表_标注者A.xlsx   难例 {len(hard_rows)} + 复核 {len(recheck_rows)} = "
          f"{len(hard_rows) + len(recheck_rows)} 条")
    print(f"  抽检表_标注者B.xlsx   主样本 {len(main_rows)} + 难例 {len(hard_rows)} = "
          f"{len(main_rows) + len(hard_rows)} 条")
    print(f"\n  难例层分布 {dict(Counter(k for _, k in hard))}"
          f"（A=grounded error, B=证据跨表页, C=带货币单位, D=比率/每股类）")
    print(f"  难例题型   {dict(Counter(q['type'] for q in hard_rows))}")
    print(f"  主样本与难例无重叠：{not (set(q['id'] for q in main_rows) & set(q['id'] for q in hard_rows))}")
    print(f"  两人共同标注的条目：难例 {len(hard_rows)} 条（一致性系数由此计算）")


if __name__ == "__main__":
    main()
